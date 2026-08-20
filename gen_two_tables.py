# -*- coding: utf-8 -*-
"""
models 审查版表格生成（兼容 LiteLLM 输入源）
============================================
数据源：
  1. LiteLLM_model_prices_and_context_window_20260813.json  （LiteLLM，价格 per-token USD）
  2. 新建 文本文档.md                                       （models.dev 底层 ai-sdk providers，价格 USD/1M）
辅助映射：
  3. vendor_model_families.csv  （11 家厂商 → 权威模型系列列表，用于提取"模型系列"）
  4. vendor_model_map.csv       （model_id → 厂商/系列/模式）
  5. model_mapping.csv           （可选人工覆盖：供应商/原始模型名 → 统一模型名/系列/厂商）

统一列结构（models sheet）：
  厂商 | 模型系列 | 模型名 | 原始模型名 | 供应商 | 供应商网址 | 模式 | 输入形式 |
  输入/输出/缓存读/缓存写 USD/1M | 输入/输出/缓存读/缓存写 CNY/1M |
  最大输入上下文 | 最大输出 | 价格最后刷新时间 | 数据来源

口径：
  - 厂商：生产厂商（中文优先）。推断顺序：模型名匹配 11 家系列 → provider 归一化 → provider 原始名
  - 模型系列：主版本族（去日期/厂商前缀/供应商前缀/模式后缀）。优先 11 家系列前缀匹配
  （Anthropic claude-{opus|sonnet|haiku}-X → claude-X）；未命中时使用规范化模型名
  - 模型名：去掉厂商/供应商前缀后完整保留
  - 供应商：provider；若供应商即生产厂商则记厂商名
  - 价格统一展示 USD/1M 与 CNY/1M；models 中 Moark 原值为 CNY，其余 models 原值为 USD；汇率 6.74
"""
import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

# 默认路径/常量（可被命令行参数覆盖）
LITELLM = "LiteLLM_model_prices_and_context_window_20260813.json"
MODELS = "新建 文本文档.md"
FAMILIES_CSV = "vendor_model_families.csv"
VMAP_CSV = "vendor_model_map.csv"
MAPPING_CSV = "model_mapping.csv"
CHANNEL_RATES_CSV = "channel_rates.csv"
OUT = "AI_API_两份初始表格_LiteLLM与models_统一格式.xlsx"
RATE = 6.74
NATIVE_CNY_SUPPLIERS = {"moark"}

_ap = argparse.ArgumentParser(description="生成 models 审查版表格（兼容 LiteLLM 输入源但不输出 LiteLLM sheet）")
_ap.add_argument("--litellm", default=LITELLM, help="LiteLLM 数据源 JSON 路径")
_ap.add_argument("--models", default=MODELS, help="models/ai-sdk 数据源 JSON 路径")
_ap.add_argument("--families", default=FAMILIES_CSV, help="vendor_model_families.csv 路径")
_ap.add_argument("--vmap", default=VMAP_CSV, help="vendor_model_map.csv 路径")
_ap.add_argument("--mapping", default=MAPPING_CSV, help="可选人工模型映射 CSV 路径")
_ap.add_argument("--channel-rates", default=CHANNEL_RATES_CSV, help="可选渠道汇率/计价单位 CSV 路径")
_ap.add_argument("--relaywatch-dir", default="", help="RelayWatch normalize_data.py 输出目录")
_ap.add_argument("--relaywatch-raw", default="", help="RelayWatch refresh_sites.py 原始快照路径")
_ap.add_argument("--relaywatch-config", default="relaywatch_sites.json", help="RelayWatch 渠道配置 JSON 路径")
_ap.add_argument("--out", default=OUT, help="输出 xlsx 路径")
_ap.add_argument("--rate", type=float, default=RATE, help="USD→CNY 汇率")
_args, _ = _ap.parse_known_args()
LITELLM, MODELS, FAMILIES_CSV, VMAP_CSV, MAPPING_CSV, CHANNEL_RATES_CSV, RELAYWATCH_DIR, RELAYWATCH_RAW, RELAYWATCH_CONFIG, OUT, RATE = (
    _args.litellm, _args.models, _args.families, _args.vmap, _args.mapping, _args.channel_rates,
    _args.relaywatch_dir, _args.relaywatch_raw, _args.relaywatch_config, _args.out, _args.rate)

# ---------------------------------------------------------------- 厂商归一化（provider 标识 -> (vendor_en, vendor_zh)）
VENDOR_MAP = {
    "openai": ("OpenAI", "OpenAI"), "gmi": (None, None),
    "google": ("Google", "Google"), "vertex_ai": (None, None), "vertex_ai-language-models": (None, None),
    "anthropic": ("Anthropic", "Anthropic"),
    "xai": ("xAI", "xAI"),
    "deepseek": ("DeepSeek", "深度求索"),
    "moonshot": ("Moonshot AI", "月之暗面"), "moonshotai": ("Moonshot AI", "月之暗面"),
    "minimax": ("MiniMax", "MiniMax"), "minimax-cn": ("MiniMax", "MiniMax"),
    "minimax-coding-plan": ("MiniMax", "MiniMax"), "minimax-cn-coding-plan": ("MiniMax", "MiniMax"),
    "dashscope": ("Alibaba Qwen", "阿里云通义"), "alibaba": ("Alibaba Qwen", "阿里云通义"),
    "alibaba-cn": ("Alibaba Qwen", "阿里云通义"), "alibaba-token-plan": ("Alibaba Qwen", "阿里云通义"),
    "alibaba-token-plan-cn": ("Alibaba Qwen", "阿里云通义"), "alibaba-coding-plan": ("Alibaba Qwen", "阿里云通义"),
    "alibaba-coding-plan-cn": ("Alibaba Qwen", "阿里云通义"),
    "byteplus": ("ByteDance Doubao", "字节豆包"), "volcengine": ("ByteDance Doubao", "字节豆包"),
    "mistral": ("Mistral AI", "Mistral AI"),
    "cohere": ("Cohere", "Cohere"),
    "zhipuai": ("Zhipu AI", "智谱"), "zhipuai-coding-plan": ("Zhipu AI", "智谱"),
    "zai": ("Zhipu AI", "智谱"), "zai-coding-plan": ("Zhipu AI", "智谱"),
    "siliconflow": ("SiliconFlow", "硅基流动"), "siliconflow-cn": ("SiliconFlow", "硅基流动"),
    "stepfun": ("StepFun", "阶跃星辰"), "stepfun-ai": ("StepFun", "阶跃星辰"),
    "stepfun-step-plan": ("StepFun", "阶跃星辰"), "stepfun-ai-step-plan": ("StepFun", "阶跃星辰"),
    "xiaomi": ("Xiaomi", "小米"), "xiaomi-token-plan-cn": ("Xiaomi", "小米"),
    "xiaomi-token-plan-ams": ("Xiaomi", "小米"), "xiaomi-token-plan-sgp": ("Xiaomi", "小米"),
    "meta": ("Meta", "Meta"), "nvidia": ("Nvidia", "Nvidia"),
    "groq": ("Groq", "Groq"), "sarvam": ("Sarvam AI", "Sarvam AI"),
    "upstage": ("Upstage", "Upstage"), "sakana": ("Sakana AI", "Sakana AI"),
    "poolside": ("Poolside", "Poolside"), "inception": ("Inception", "Inception"),
    "perplexity": ("Perplexity", "Perplexity"), "perplexity-agent": ("Perplexity", "Perplexity"),
    "cerebras": ("Cerebras", "Cerebras"), "friendli": ("Friendli AI", "Friendli AI"),
    "tencent-tokenhub": ("Tencent", "腾讯"), "tencent-token-plan": ("Tencent", "腾讯"),
    "tencent-coding-plan": ("Tencent", "腾讯"), "tencent-tokenhub": ("Tencent", "腾讯"),
    # 中转/云平台（供应商≠厂商）
    "azure": (None, None), "azure_ai": (None, None), "azure-ai": (None, None),
    "bedrock": (None, None), "bedrock_converse": (None, None), "amazon-bedrock": (None, None),
    "openrouter": (None, None), "vercel": (None, None), "vercel_ai_gateway": (None, None),
    "together_ai": (None, None), "togetherai": (None, None),
    "fireworks_ai": (None, None), "fireworks-ai": (None, None),
    "deepinfra": (None, None), "novita": (None, None), "novita-ai": (None, None),
    "replicate": (None, None), "huggingface": (None, None),
    "groq": ("Groq", "Groq"),
}

def norm(s):
    return str(s).lower().replace("_", "-").replace(":", "-")

# ---------------------------------------------------------------- 加载 families（11 家权威系列）
families = {}          # vendor_en -> (vendor_zh, [family])
with open(FAMILIES_CSV, encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        fams = [x.strip() for x in row["model_families"].split("|") if x.strip()]
        families[row["vendor_en"]] = (row["vendor_zh"], fams)

# 全量 family 前缀索引（跨厂商，用于推断生产厂商）
fam_index = []         # (norm_family, vendor_en, vendor_zh, family)
for ven, (vzh, fams) in families.items():
    for fa in fams:
        fam_index.append((norm(fa), ven, vzh, fa))
fam_index.sort(key=lambda x: -len(x[0]))

# ---------------------------------------------------------------- 加载 vendor_model_map（model_id -> mode）
vmap = {}
with open(VMAP_CSV, encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        vmap[row["model_id"]] = row["mode"]
vmap_norm = {norm(k): v for k, v in vmap.items()}

# 可选人工映射：用于非头部中转站将原始模型名明确归一到统一模型名/系列/厂商。
# 匹配优先级为“供应商 + 原始模型名”，再回退到“任意供应商 + 原始模型名”。
manual_model_map = {}
try:
    with open(MAPPING_CSV, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            raw = (row.get("raw_model_name") or row.get("原始模型名") or "").strip()
            if not raw:
                continue
            supplier = (row.get("supplier") or row.get("供应商") or "").strip()
            manual_model_map[(norm(supplier), norm(raw))] = {
                "model": (row.get("model_name") or row.get("统一模型名") or "").strip() or None,
                "series": (row.get("model_series") or row.get("模型系列") or "").strip() or None,
                "vendor": (row.get("vendor") or row.get("厂商") or "").strip() or None,
            }
except FileNotFoundError:
    pass

# RelayWatch 渠道配置由网站设置持久化，包含网址、显示名、汇率和计价单位。
relaywatch_config_map = {}
try:
    relaywatch_config_rows = json.loads(Path(RELAYWATCH_CONFIG).read_text(encoding="utf-8"))
    for item in relaywatch_config_rows if isinstance(relaywatch_config_rows, list) else []:
        if not isinstance(item, dict):
            continue
        for key in (item.get("url"), item.get("origin"), item.get("pricingUrl")):
            if key:
                relaywatch_config_map[str(key).rstrip("/").lower()] = item
except (FileNotFoundError, json.JSONDecodeError):
    pass

def manual_mapping(provider, provider_name, raw_model_name):
    raw_key = norm(raw_model_name)
    return manual_model_map.get((norm(provider_name), raw_key)) or manual_model_map.get((norm(provider), raw_key)) or manual_model_map.get(("", raw_key))

# 渠道级换算设置。默认按 USD 计价，汇率为 CNY / 计价单位；将计价单位改为“代币”
# 后，同样的数字会按代币折合人民币，不再把它当作真实美元。
channel_rate_map = {}
try:
    with open(CHANNEL_RATES_CSV, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            key = (row.get("supplier") or row.get("渠道(供应商)") or "").strip()
            if not key:
                continue
            raw_rate = (row.get("exchange_rate") or row.get("汇率(CNY/计价单位)") or "").strip()
            try:
                exchange_rate = float(raw_rate) if raw_rate else RATE
            except ValueError:
                exchange_rate = RATE
            unit = (row.get("billing_unit") or row.get("计价单位") or "USD").strip() or "USD"
            note = (row.get("note") or row.get("备注") or "").strip()
            pricing_mode = (row.get("pricing_mode") or row.get("价格解析方式") or "auto").strip().lower() or "auto"
            try:
                price_scale = float(row.get("price_scale") or row.get("额外价格倍率") or 1)
            except (TypeError, ValueError):
                price_scale = 1.0
            if price_scale <= 0:
                price_scale = 1.0
            channel_rate_map[norm(key)] = {"rate": exchange_rate, "unit": unit, "note": note,
                                           "pricing_mode": pricing_mode, "price_scale": price_scale}
except FileNotFoundError:
    pass

def channel_settings(*names):
    for name in names:
        if name and norm(name) in channel_rate_map:
            return channel_rate_map[norm(name)]
    return {"rate": RATE, "unit": "USD", "note": "", "pricing_mode": "auto", "price_scale": 1.0}

def is_usd_unit(unit):
    return norm(unit) in {"usd", "美元", "$", "us-dollar", "dollar"}

def convert_native_prices(values, unit, exchange_rate, usd_exchange_rate=RATE):
    """将渠道原始计价单位换算为 USD/CNY。

    ``exchange_rate`` 是渠道配置的 CNY/计价单位；``usd_exchange_rate``
    是渠道自己公布的 USD/CNY 汇率，适用于人民币渠道，避免使用全局
    汇率导致不同渠道的 USD 参考价漂移。
    """
    usd_values, cny_values = [], []
    unit_norm = norm(unit)
    try:
        usd_rate = float(usd_exchange_rate)
    except (TypeError, ValueError):
        usd_rate = RATE
    if usd_rate <= 0:
        usd_rate = RATE
    for value in values:
        if value is None:
            usd_values.append(None)
            cny_values.append(None)
            continue
        try:
            native = float(value)
        except (TypeError, ValueError):
            usd_values.append(None)
            cny_values.append(None)
            continue
        if is_usd_unit(unit):
            usd = native
            cny = native * exchange_rate
        elif unit_norm in {"cny", "人民币", "rmb", "元"}:
            cny = native
            usd = native / usd_rate if usd_rate else None
        else:
            cny = native * exchange_rate
            usd = cny / usd_rate if usd_rate else None
        usd_values.append(round(usd, 6) if usd is not None else None)
        cny_values.append(round(cny, 6) if cny is not None else None)
    return usd_values, cny_values

def match_family(model_lower, fams):
    best = None
    for fa in fams:
        fl = norm(fa)
        if model_lower == fl or model_lower.startswith(fl + "-") or model_lower.startswith(fl + "."):
            if best is None or len(fl) > len(best):
                best = fa
    return best

def extract_family(model_id):
    """返回 (family, vendor_en, vendor_zh)；未命中返回 (None, None, None)"""
    ml = norm(model_id)
    for fl, ven, vzh, fa in fam_index:
        if ml == fl or ml.startswith(fl + "-") or ml.startswith(fl + "."):
            return fa, ven, vzh
    # Anthropic 档位跳过：claude-{opus|sonnet|haiku}-X -> claude-X
    m = re.match(r"^claude-(opus|sonnet|haiku)-(.+)$", ml)
    if m:
        cand = "claude-" + m.group(2)
        for fl, ven, vzh, fa in fam_index:
            if ven == "Anthropic" and (cand == fl or cand.startswith(fl + "-")):
                return fa, ven, vzh
    return None, None, None

def vendor_of(provider):
    """provider 标识 -> (vendor_en, vendor_zh)，未知返回 (None, None)"""
    return VENDOR_MAP.get(provider, VENDOR_MAP.get(norm(provider), (None, None)))

# 厂商关键词（直接从模型名提取生产厂商，用户要求的最直接方式）
# 正则词边界匹配，避免短词误命中
VENDOR_KEYWORDS = [
    ("OpenAI", "OpenAI", [r"\bchatgpt\b", r"\bgpt\b", r"\bdall-e", r"\bwhisper\b", r"\bcodex\b", r"\bsora\b", r"\bo1\b", r"\bo3\b", r"\bo4\b"]),
    ("Anthropic", "Anthropic", [r"\bclaude\b"]),
    ("Google", "Google", [r"\bgemini\b", r"\bgemma\b", r"\bpalm\b", r"\bimagen\b", r"\bveo\b", r"\blearnlm\b", r"\bmedlm\b", r"\blyria\b", r"\bchirp\b"]),
    ("xAI", "xAI", [r"\bgrok\b"]),
    ("DeepSeek", "深度求索", [r"\bdeepseek\b"]),
    ("Moonshot AI", "月之暗面", [r"\bkimi\b", r"\bmoonshot\b"]),
    ("MiniMax", "MiniMax", [r"\bminimax\b", r"\babab\b"]),
    ("Alibaba Qwen", "阿里云通义", [r"\bqwen\b"]),
    ("ByteDance Doubao", "字节豆包", [r"\bdoubao\b"]),
    ("Xiaomi", "小米", [r"\bmimo\b"]),
    ("Zhipu AI", "智谱", [r"\bglm\b", r"\bchatglm\b", r"\bcogview\b"]),
    ("Mistral AI", "Mistral AI", [r"\bmistral\b", r"\bmixtral\b", r"\bcodestral\b", r"\bpixtral\b", r"\bministral\b", r"\bmagistral\b", r"\bvoxtral\b", r"\bdevstral\b"]),
    ("Cohere", "Cohere", [r"\bcohere\b", r"\bcommand\b", r"\baya\b"]),
    ("Meta", "Meta", [r"\bllama\b"]),
    ("Stability AI", "Stability AI", [r"\bstable-diffusion\b", r"\bsdxl\b", r"\bsd3\b"]),
    ("Black Forest Labs", "Black Forest Labs", [r"\bflux\b"]),
    ("Tencent", "腾讯", [r"\bhunyuan\b"]),
    ("Baidu", "百度", [r"\bernie\b"]),
    ("Nvidia", "Nvidia", [r"\bnemotron\b"]),
    ("01.AI", "零一万物", [r"\byi-lightning\b", r"\byi-vision\b", r"\byi-\d"]),
    ("StepFun", "阶跃星辰", [r"\bstep-?\d"]),
    ("Amazon", "Amazon", [r"\bnova-\d", r"\btitan\b"]),
    ("Microsoft", "微软", [r"\bphi-\d"]),
    ("AI21", "AI21", [r"\bjamba\b"]),
    ("Snowflake", "Snowflake", [r"\barctic\b"]),
    ("Sakana AI", "Sakana AI", [r"\bsakana\b"]),
    ("Poolside", "Poolside", [r"\bpoolside\b", r"\bptx\b"]),
    ("Upstage", "Upstage", [r"\bsolar\b"]),
    ("Sarvam AI", "Sarvam AI", [r"\bsarvam\b"]),
]

def vendor_by_keyword(model_name):
    """模型名关键词 -> 生产厂商（最直接判断）"""
    ml = norm(model_name)
    for ven, vzh, pats in VENDOR_KEYWORDS:
        for p in pats:
            if re.search(p, ml):
                return ven, vzh
    return None, None

def resolve_vendor(model_name, provider, provider_name=None):
    """厂商推断优先级：模型名关键词 > 11 家系列匹配 > provider 归一化 > provider 名"""
    kven, kvzh = vendor_by_keyword(model_name)
    if kven:
        return kven, kvzh
    _, ven, vzh = extract_family(model_name)
    if ven:
        return ven, vzh
    pven, pvzh = vendor_of(provider)
    if pven:
        return pven, pvzh
    return (provider_name or provider), (provider_name or provider)

# 已知渠道官网/定价页（LiteLLM 侧 source 缺失时的补充；公开事实）
PROVIDER_URL = {
    "openai": "https://platform.openai.com/docs/pricing",
    "gemini": "https://ai.google.dev/gemini-api/docs/pricing",
    "anthropic": "https://docs.anthropic.com/en/docs/about-claude/pricing-overview",
    "deepseek": "https://api-docs.deepseek.com/quick_start/pricing",
    "dashscope": "https://www.alibabacloud.com/help/en/model-studio/models",
    "moonshot": "https://platform.moonshot.ai/docs/pricing/chat",
    "minimax": "https://www.minimax.io/pricing",
    "xai": "https://docs.x.ai/docs/models",
    "mistral": "https://mistral.ai/pricing",
    "cohere": "https://cohere.com/pricing",
    "groq": "https://groq.com/pricing",
    "perplexity": "https://docs.perplexity.ai/guides/pricing",
    "together_ai": "https://www.together.ai/pricing",
    "fireworks_ai": "https://fireworks.ai/pricing",
    "deepinfra": "https://deepinfra.com/pricing",
    "novita": "https://novita.ai/pricing",
    "openrouter": "https://openrouter.ai/models",
    "bedrock": "https://aws.amazon.com/bedrock/pricing/",
    "bedrock_converse": "https://aws.amazon.com/bedrock/pricing/",
    "azure": "https://azure.microsoft.com/en-us/pricing/details/cognitive-services/openai-service/",
    "azure_ai": "https://azure.microsoft.com/en-us/pricing/details/cognitive-services/openai-service/",
    "vertex_ai": "https://cloud.google.com/vertex-ai/generative-ai/pricing",
    "vertex_ai-language-models": "https://cloud.google.com/vertex-ai/generative-ai/pricing",
    "replicate": "https://replicate.com/pricing",
    "huggingface": "https://huggingface.co/pricing",
    "vercel_ai_gateway": "https://vercel.com/docs/ai-gateway",
    "snowflake": "https://www.snowflake.com/en/data-cloud/cortex/",
    "databricks": "https://www.databricks.com/product/pricing",
    "cerebras": "https://cerebras.ai/pricing",
    "friendli": "https://friendli.ai/pricing",
    "oci": "https://www.oracle.com/cloud/ai/generative-ai/pricing/",
    "watsonx": "https://www.ibm.com/products/watsonx-ai/pricing",
    "github": "https://docs.github.com/en/copilot/",
    "nvidia": "https://build.nvidia.com/pricing",
}

# 厂商/供应商/区域前缀黑名单（模型名开头需删除的冗余前缀）
PREFIX_BLACKLIST = {
    # 区域
    "us", "eu", "apac", "ap", "na", "sa", "cn", "ams", "sgp",
    # 官方厂商（作为冗余前缀出现时删：anthropic.claude-...、openai.gpt-...、qwen.qwen3-...）
    "anthropic", "openai", "google", "minimax", "meta", "amazon", "cohere", "qwen",
    "moonshotai", "moonshot", "stability", "xiaomi", "mistral",
    # 供应商（databricks-gpt-5、lucidquery-agi-...、alicloud-deepseek-v4、umans-...）
    "databricks", "lucidquery", "lucidnova", "alicloud", "umans", "gradient",
}
# 模型系列开头提示：遇到这些开头说明已经到真正的模型名，停止删前缀
MODEL_SERIES_HINTS = {
    "gpt", "chatgpt", "claude", "gemini", "grok", "deepseek", "kimi", "moonshot",
    "qwen", "doubao", "glm", "llama", "mistral", "mixtral", "codestral", "pixtral",
    "ministral", "magistral", "voxtral", "devstral", "minimax", "mimo", "flux",
    "gemma", "palm", "imagen", "veo", "learnlm", "medlm", "lyria", "chirp",
    "jamba", "command", "aya", "phi", "nemotron", "nova", "titan", "stable",
    "dall", "whisper", "sora", "o1", "o3", "o4", "hunyuan", "ernie", "yi", "step",
    "solar", "sarvam", "arctic", "ptx", "sakana", "poolside", "codestral", "ocr",
    "sonar", "perplexity", "text", "tts", "embedding", "embed", "rerank", "image",
    "audio", "video", "moderation", "nano", "green", "holo2", "muse", "hy3", "mimo",
}

_PREFIX_RE = re.compile(
    r"^(?:" + "|".join(sorted(PREFIX_BLACKLIST, key=len, reverse=True)) + r")\.")

def normalize_model_name(model_id):
    """统一模型名：去斜杠前缀 + 删厂商/供应商/区域前缀 + 小写 + p→点号 + 下划线/冒号转连字符"""
    ml = str(model_id).strip()
    if "/" in ml:
        ml = ml.rsplit("/", 1)[-1]
    ml = ml.lower()
    ml = ml.replace("_", "-").replace(":", "-").replace(" ", "-")
    # 数字p数字 -> 数字.数字（qwen2p5 -> qwen2.5、m2p7 -> m2.7、3p5 -> 3.5）
    ml = re.sub(r"(\d)p(\d)", r"\1.\2", ml)
    # deep-deepseek 去重前缀
    ml = re.sub(r"^deep-deepseek", "deepseek", ml)
    # 1) 删点号前缀段（us. / anthropic. / minimax. / meta. / qwen. 等），循环
    while True:
        new = _PREFIX_RE.sub("", ml, count=1)
        if new == ml:
            break
        ml = new
    # 2) 删连字符首段（databricks- / alicloud- / umans- / openai- 等，且不是模型系列）
    m = re.match(r"^([a-z0-9]+)-(.*)$", ml)
    if m and m.group(1) in PREFIX_BLACKLIST and m.group(1) not in MODEL_SERIES_HINTS:
        ml = m.group(2)
    ml = re.sub(r"-+", "-", ml).strip("-")
    return ml

MODE_HINTS = [
    ("embedding", "embedding"), ("embed", "embedding"),
    ("rerank", "rerank"), ("image", "image_generation"), ("vision", "chat"),
    ("realtime", "realtime"), ("audio", "audio_speech"), ("transcribe", "audio_transcription"),
    ("tts", "audio_speech"), ("moderation", "moderation"), ("video", "video_generation"),
    ("ocr", "ocr"),
]
def infer_mode(model_id):
    ml = norm(model_id)
    if ml in vmap_norm:
        return vmap_norm[ml]
    for kw, mode in MODE_HINTS:
        if kw in ml:
            return mode
    return None

def format_input_modalities(modalities):
    """保留 models.dev 的 input 形式，便于表格和网站展示。"""
    if isinstance(modalities, dict):
        values = modalities.get("input")
    else:
        values = modalities
    if isinstance(values, (list, tuple)):
        return ", ".join(str(value).strip() for value in values if str(value).strip()) or None
    if values is None:
        return None
    return str(values).strip() or None

def fmt(v):
    """将价格嵌入 Excel 公式时保持紧凑的小数表示。"""
    if v is None:
        return "-"
    s = f"{float(v):.6f}".rstrip("0").rstrip(".")
    return s or "0"

# 模式/档位后缀白名单（用于 fallback 提取主版本族）
MODE_SUFFIXES = [
    "highspeed", "thinking", "reasoning", "preview", "instruct", "latest",
    "research", "search", "vision", "audio", "image", "video", "flashx", "flash",
    "pro", "mini", "nano", "lite", "turbo", "plus", "high", "low", "fast", "speed",
    "air", "code", "coder", "tts", "opus", "sonnet", "haiku", "vl", "omni", "multi",
]
def strip_mode_suffix(name):
    """去掉日期与模式/档位后缀，得到主版本族（fallback 用）"""
    s = str(name)
    s = re.sub(r"-\d{8}$", "", s)
    s = re.sub(r"-\d{4}[-.]\d{2}[-.]\d{2}.*$", "", s)
    changed = True
    while changed:
        changed = False
        for sfx in MODE_SUFFIXES:
            if s.lower().endswith("-" + sfx):
                s = s[:-(len(sfx) + 1)]
                changed = True
    return s

# ---------------------------------------------------------------- 兼容读取 LiteLLM 源（不写入审查版工作簿）
with open(LITELLM, encoding="utf-8") as f:
    lit = json.load(f)
lit_real = {k: v for k, v in lit.items() if k != "sample_spec"}

def to_1m(x):
    return round(x * 1_000_000, 6) if x is not None else None

def cny(x):
    return round(x * RATE, 6) if x is not None else None

def usd_from_cny(x):
    return round(x / RATE, 6) if x is not None else None

lit_rows = []
lit_miss_fam = 0
for key, v in lit_real.items():
    provider = v.get("litellm_provider") or ""
    manual = manual_mapping(provider, provider, key)
    model_name = manual.get("model") if manual else None
    model_name = model_name or normalize_model_name(key)
    if not model_name:
        continue   # 过滤目录项（如 fireworks_ai/accounts/fireworks/models/）
    fam, ven, vzh = extract_family(model_name)
    if fam is None:
        lit_miss_fam += 1
        fam = strip_mode_suffix(model_name) or model_name   # fallback：去日期/档位后缀
    if manual:
        fam = manual.get("series") or fam
        if manual.get("vendor"):
            ven = vzh = manual["vendor"]
    else:
        ven, vzh = resolve_vendor(model_name, provider)
    supplier = vzh if (vendor_of(provider)[0] == ven) else provider
    mode = v.get("mode") or infer_mode(model_name)
    usd_in = to_1m(v.get("input_cost_per_token"))
    usd_out = to_1m(v.get("output_cost_per_token"))
    usd_cr = to_1m(v.get("cache_read_input_token_cost"))
    usd_cw = to_1m(v.get("cache_creation_input_token_cost"))
    ctx = v.get("max_input_tokens") or v.get("max_tokens")
    out_tok = v.get("max_output_tokens")
    lit_rows.append([
        vzh, fam, model_name, key, supplier,
        v.get("source") or PROVIDER_URL.get(provider, PROVIDER_URL.get(norm(provider))), mode,
        usd_in, usd_out, usd_cr, usd_cw,
        cny(usd_in), cny(usd_out), cny(usd_cr), cny(usd_cw),
        int(ctx) if ctx is not None else None,
        int(out_tok) if out_tok is not None else None,
        "litellm",
    ])

# ---------------------------------------------------------------- 生成 models 表
with open(MODELS, encoding="utf-8") as f:
    md = json.load(f)

# 定价页映射（供应商显示名 -> 定价页 URL，来自全量抓取结果；缺失则回退 doc）
PRICING = {}
try:
    with open("pricing_pages.json", encoding="utf-8") as f:
        PRICING = {v["name"]: v.get("pricing") for v in json.load(f).values() if v.get("pricing")}
except Exception:
    PRICING = {}

def pricing_url_of(prov_name, doc):
    """供应商网址：优先定价页，缺失回退 doc"""
    return PRICING.get(prov_name) or doc
models_rows = []
models_miss_fam = 0
for prov, pv in md.items():
    prov_name = pv.get("name") or prov
    is_native_cny = norm(prov) in NATIVE_CNY_SUPPLIERS or norm(prov_name) in NATIVE_CNY_SUPPLIERS
    channel_cfg = channel_settings(prov_name, prov)
    channel_rate = channel_cfg["rate"]
    has_manual_unit = any(norm(name) in channel_rate_map for name in (prov_name, prov) if name)
    channel_unit = channel_cfg["unit"] if has_manual_unit else ("CNY" if is_native_cny else "USD")
    pven, pvzh = vendor_of(prov)
    for mid, mv in pv.get("models", {}).items():
        manual = manual_mapping(prov, prov_name, mid)
        clean_name = (manual.get("model") if manual else None) or normalize_model_name(mid)       # 统一模型名：删前缀 + 小写 + p→点号
        fam, ven, vzh = extract_family(clean_name)
        if fam is None:
            models_miss_fam += 1
            fam = strip_mode_suffix(clean_name) or clean_name   # fallback：去日期/档位后缀
        if manual:
            fam = manual.get("series") or fam
            if manual.get("vendor"):
                ven = vzh = manual["vendor"]
            else:
                ven, vzh = resolve_vendor(clean_name, prov, prov_name)
        else:
            ven, vzh = resolve_vendor(clean_name, prov, prov_name)
        supplier = vzh if (vendor_of(prov)[0] == ven) else prov_name
        cost = mv.get("cost", {}) or {}
        lim = mv.get("limit", {}) or {}
        is_plan = "plan" in prov.lower() or "plan" in prov_name.lower()
        raw_in = cost.get("input")
        raw_out = cost.get("output")
        raw_cr = cost.get("cache_read")
        raw_cw = cost.get("cache_write")
        if is_plan:
            # 套餐计划（Coding/Token Plan）的 0 表示"含在套餐内"，非 0 元单价 → 视为无公开单价
            raw_in = None if raw_in == 0 else raw_in
            raw_out = None if raw_out == 0 else raw_out
            raw_cr = None if raw_cr == 0 else raw_cr
            raw_cw = None if raw_cw == 0 else raw_cw
        usd_values, cny_values = convert_native_prices(
            (raw_in, raw_out, raw_cr, raw_cw), channel_unit, channel_rate
        )
        usd_in, usd_out, usd_cr, usd_cw = usd_values
        cny_in, cny_out, cny_cr, cny_cw = cny_values
        ctx = lim.get("context") or lim.get("input")
        out_tok = lim.get("output")
        mode = vmap_norm.get(norm(clean_name)) or infer_mode(clean_name)
        models_rows.append([
            vzh, fam, clean_name, mid, supplier, pricing_url_of(prov_name, pv.get("doc")), mode,
            format_input_modalities(mv.get("modalities")),
            usd_in, usd_out, usd_cr, usd_cw,
            cny_in, cny_out, cny_cr, cny_cw,
            int(ctx) if ctx is not None else None,
            int(out_tok) if out_tok is not None else None,
            mv.get("last_updated") or mv.get("lastUpdated"),
            "models.dev",
            channel_rate,
            channel_cfg["unit"],
            channel_cfg.get("pricing_mode", "model_ratio"),
            channel_cfg.get("price_scale", 1.0),
        ])

# ---------------------------------------------------------------- 合并 RelayWatch 中转站数据
# RelayWatch 已负责 NewAPI/Sub2API 端点发现、归一化和倍率读取；这里仅把其
# 归一化结果映射到当前工作簿的 USD/CNY 每百万 token 结构。
relaywatch_channels = []
relaywatch_row_count = 0
if RELAYWATCH_DIR:
    relay_dir = Path(RELAYWATCH_DIR)
    try:
        relay_sites = json.loads((relay_dir / "sites.json").read_text(encoding="utf-8"))
        relay_models = json.loads((relay_dir / "models.json").read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        relay_sites, relay_models = [], []
    raw_by_origin = {}
    if RELAYWATCH_RAW:
        try:
            raw_payload = json.loads(Path(RELAYWATCH_RAW).read_text(encoding="utf-8"))
            for raw_row in raw_payload.get("rows", raw_payload if isinstance(raw_payload, list) else []):
                raw_by_origin[raw_row.get("origin")] = raw_row
        except (FileNotFoundError, json.JSONDecodeError):
            raw_by_origin = {}
    site_by_origin = {site.get("origin"): site for site in relay_sites if site.get("origin")}

    def relay_raw_config(origin):
        raw_row = raw_by_origin.get(origin) or {}
        status_body = ((raw_row.get("endpoints") or {}).get("status") or {}).get("body") or ""
        try:
            parsed = json.loads(status_body)
        except (TypeError, json.JSONDecodeError):
            parsed = {}
        data = parsed.get("data") if isinstance(parsed, dict) else {}
        return data if isinstance(data, dict) else {}

    def relay_price(values, unit, rate, usd_exchange_rate=RATE):
        if values is None:
            return None, None
        try:
            native = float(values)
        except (TypeError, ValueError):
            return None, None
        usd_values, cny_values = convert_native_prices((native,), unit, rate, usd_exchange_rate)
        return usd_values[0], cny_values[0]

    for site in relay_sites:
        origin = site.get("origin") or ""
        if not origin:
            continue
        raw_row = raw_by_origin.get(origin) or {}
        requested_origin = raw_row.get("requested_origin") or origin
        relay_site_config = (
            relaywatch_config_map.get(origin.rstrip("/").lower())
            or relaywatch_config_map.get(str(requested_origin).rstrip("/").lower())
            or {}
        )
        site_name = str(relay_site_config.get("name") or site.get("name") or origin.split("//", 1)[-1].split(".")[0]).strip()
        cfg = relay_raw_config(origin)
        explicit_cfg = channel_settings(site_name, requested_origin, origin)
        has_manual_rate = any(norm(name) in channel_rate_map for name in (site_name, requested_origin, origin) if name)
        raw_rate = cfg.get("usd_exchange_rate") or cfg.get("custom_currency_exchange_rate")
        try:
            raw_rate = float(raw_rate)
        except (TypeError, ValueError):
            raw_rate = RATE
        configured_rate = relay_site_config.get("exchangeRate")
        try:
            configured_rate = float(configured_rate) if configured_rate is not None else None
        except (TypeError, ValueError):
            configured_rate = None
        rate = configured_rate or (explicit_cfg["rate"] if has_manual_rate else raw_rate)
        unit_from_cfg = str(relay_site_config.get("billingUnit") or "").strip() or (explicit_cfg["unit"] if has_manual_rate else "")
        quota_type = str(cfg.get("quota_display_type") or "").upper()
        custom_symbol = str(cfg.get("custom_currency_symbol") or "").strip()
        unit = unit_from_cfg or ("USD" if quota_type == "USD" or custom_symbol == "$" else "CNY" if quota_type == "CNY" or custom_symbol in {"¥", "￥"} else custom_symbol or "代币")
        pricing_mode = str(relay_site_config.get("pricingMode") or explicit_cfg.get("pricing_mode") or "auto").strip().lower()
        if pricing_mode not in {"auto", "billing_expr", "model_ratio"}:
            pricing_mode = "auto"
        try:
            price_scale = float(relay_site_config.get("priceScale", explicit_cfg.get("price_scale", 1.0)))
        except (TypeError, ValueError):
            price_scale = 1.0
        if price_scale <= 0:
            price_scale = 1.0
        try:
            raw_usd_exchange_rate = float(raw_rate)
        except (TypeError, ValueError):
            raw_usd_exchange_rate = RATE
        if raw_usd_exchange_rate <= 0:
            raw_usd_exchange_rate = RATE
        pricing_url = str(relay_site_config.get("pricingUrl") or requested_origin).strip()
        api_base = str(relay_site_config.get("apiBase") or (origin.rstrip("/") + "/v1")).strip()
        models_for_site = []
        for model_record in relay_models:
            for model_site in model_record.get("sites", []) or []:
                if model_site.get("origin") != origin:
                    continue
                raw_name = str(model_site.get("raw_model") or model_site.get("model") or "").strip()
                clean_name = normalize_model_name(raw_name)
                if not clean_name:
                    continue
                fam, ven, vzh = extract_family(clean_name)
                if fam is None:
                    fam = strip_mode_suffix(clean_name) or clean_name
                provider = str(model_site.get("provider") or "").strip()
                ven, vzh = resolve_vendor(clean_name, provider or site_name, provider or site_name)
                mode = infer_mode(clean_name) or ("image_generation" if "image" in clean_name else "chat")
                modalities = "text, image" if any(token in clean_name for token in ("image", "vision")) else "text"
                factor = model_site.get("token_price_multiplier")
                try:
                    factor = float(factor)
                except (TypeError, ValueError):
                    factor = 1_000_000 / float(cfg.get("quota_per_unit") or 500_000)
                ratio = model_site.get("model_ratio")
                completion = model_site.get("completion_ratio")
                cache_read = model_site.get("cache_ratio")
                cache_write = model_site.get("create_cache_ratio")
                coefficients = model_site.get("billing_expr_coefficients") or {}
                use_expression = pricing_mode == "billing_expr" or (pricing_mode == "auto" and coefficients.get("input") is not None)
                if use_expression and coefficients.get("input") is not None:
                    # Billing expressions are expressed in the provider's
                    # quota currency. For CNY displays, ESEN-style services
                    # multiply the coefficients by their published unit
                    # price (7.3 in the current snapshot); USD expressions
                    # already represent USD per million tokens.
                    unit_norm = norm(unit)
                    expression_unit_price = model_site.get("currency_unit_price")
                    try:
                        expression_unit_price = float(expression_unit_price)
                    except (TypeError, ValueError):
                        expression_unit_price = 1.0
                    expression_factor = expression_unit_price if unit_norm in {"cny", "人民币", "rmb", "元"} else 1.0
                    native_values = [
                        (coefficients.get(key) * expression_factor * price_scale
                         if coefficients.get(key) is not None else None)
                        for key in ("input", "output", "cache_read", "cache_write")
                    ]
                else:
                    native_values = []
                    for multiplier in (1, completion, cache_read, cache_write):
                        try:
                            native_values.append(float(ratio) * float(multiplier) * factor * price_scale if ratio is not None and multiplier is not None else None)
                        except (TypeError, ValueError):
                            native_values.append(None)
                usd_values, cny_values = [], []
                for native in native_values:
                    usd_value, cny_value = relay_price(native, unit, rate, raw_usd_exchange_rate)
                    usd_values.append(usd_value)
                    cny_values.append(cny_value)
                if all(value is None for value in usd_values) and all(value is None for value in cny_values):
                    continue
                models_rows.append([
                    vzh, fam, clean_name, raw_name, site_name, pricing_url, mode, modalities,
                    *usd_values, *cny_values,
                    None, None, raw_row.get("scanned_at") or site.get("updated_at"), "relaywatch爬取",
                    rate, unit, pricing_mode, price_scale,
                ])
                models_for_site.append(clean_name)
                relaywatch_row_count += 1
        relaywatch_channels.append({
            "name": site_name,
            "vendor": "中转/云平台",
            "source": "relaywatch爬取",
            "model_count": len(set(models_for_site)),
            "pricing_url": pricing_url,
            "api": api_base,
            "rate": rate,
            "unit": unit,
            "pricing_mode": pricing_mode,
            "price_scale": price_scale,
        })

# ---------------------------------------------------------------- 写出
HEADER = ["厂商", "模型系列", "模型名", "原始模型名", "供应商", "供应商网址", "模式", "输入形式",
          "输入价格USD/1M", "输出价格USD/1M", "缓存读取USD/1M", "缓存写入USD/1M",
          "输入价格CNY/1M", "输出价格CNY/1M", "缓存读取CNY/1M", "缓存写入CNY/1M",
          "最大输入上下文", "最大输出", "价格最后刷新时间", "数据来源",
          "渠道汇率(CNY/计价单位)", "计价单位", "价格解析方式", "额外价格倍率"]

def safe(v):
    if isinstance(v, str) and v.startswith(("=", "+", "-", "@")):
        return "'" + v
    return v

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
wb = openpyxl.Workbook()
wb.remove(wb.active)
# 审查版只保留 models 数据；LiteLLM 仍可作为输入源，但不再写入工作簿。
for sheet, rows in [("models", models_rows)]:
    ws = wb.create_sheet(sheet)
    ws.append([safe(x) for x in HEADER])
    for r in rows:
        ws.append([safe(x) for x in r])
    thin = openpyxl.styles.Side(style="thin", color="BFBFBF")
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = openpyxl.styles.PatternFill("solid", fgColor="D9E2F3")
    for c in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = fill
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADER)):
        for c in row:
            c.border = border
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 24
    for col in "IJKLMNOP":
        ws.column_dimensions[col].width = 15
    for col in ["Q", "R"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["S"].width = 18
    ws.column_dimensions["T"].width = 12
    ws.column_dimensions["U"].width = 18
    ws.column_dimensions["V"].width = 12
    ws.auto_filter.ref = f"A1:V{ws.max_row}"
    ws.freeze_panes = "D2"

# ---------------------------------------------------------------- 横向对比表（由 models 数据生成；每个模型 4 行并合并模型/厂商单元格）
COMPARE_TYPES = ["输入", "输出", "缓存读", "缓存写"]
compare_models = {}
compare_suppliers = set()
for model_row in models_rows:
    model_name = str(model_row[2] or "").strip()
    supplier = str(model_row[4] or "").strip()
    if not model_name or not supplier:
        continue
    usd_vals = model_row[8:12]
    cny_vals = model_row[12:16]
    if all(v is None for v in usd_vals) and all(v is None for v in cny_vals):
        continue
    compare_suppliers.add(supplier)
    rec = compare_models.setdefault(model_name, {"vendor": model_row[0] or "其他", "prices": {}})
    rec["prices"][supplier] = (usd_vals, cny_vals)

compare_suppliers = sorted(compare_suppliers)
compare_model_items = sorted(compare_models.items())
ws_cmp = wb.create_sheet("横向对比")
ws_cmp.cell(row=1, column=1, value="模型 × 供应商 定价对比表（models 数据，每模型纵向 4 行）")
ws_cmp.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
ws_cmp.cell(row=2, column=1, value="币种切换:")
ws_cmp.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)
ws_cmp.cell(row=2, column=2, value="USD")
ws_cmp.cell(row=2, column=3, value="单位：每百万 token；选择 USD/CNY 切换显示")
ws_cmp.cell(row=2, column=3).font = openpyxl.styles.Font(size=9, color="808080")
cmp_dv = DataValidation(type="list", formula1='"USD,CNY"', allow_blank=False)
ws_cmp.add_data_validation(cmp_dv)
cmp_dv.add(ws_cmp.cell(row=2, column=2))

cmp_header_fill = openpyxl.styles.PatternFill("solid", fgColor="D9E2F3")
cmp_type_fill = openpyxl.styles.PatternFill("solid", fgColor="E2EFDA")
ws_cmp.cell(row=4, column=1, value="模型")
ws_cmp.cell(row=4, column=2, value="厂商")
ws_cmp.cell(row=4, column=3, value="类型")
for c in range(1, 4):
    ws_cmp.cell(row=4, column=c).font = openpyxl.styles.Font(bold=True)
    ws_cmp.cell(row=4, column=c).fill = cmp_header_fill
for j, supplier in enumerate(compare_suppliers, start=4):
    cell = ws_cmp.cell(row=4, column=j, value=supplier)
    cell.font = openpyxl.styles.Font(bold=True, size=9)
    cell.fill = cmp_header_fill
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

cmp_row = 5
for model_name, rec in compare_model_items:
    block_start = cmp_row
    for type_index, type_label in enumerate(COMPARE_TYPES):
        if type_index == 0:
            ws_cmp.cell(row=cmp_row, column=1, value=model_name)
            ws_cmp.cell(row=cmp_row, column=2, value=rec["vendor"])
        ws_cmp.cell(row=cmp_row, column=3, value=type_label)
        ws_cmp.cell(row=cmp_row, column=3).fill = cmp_type_fill
        ws_cmp.cell(row=cmp_row, column=3).font = openpyxl.styles.Font(bold=True, size=9)
        ws_cmp.cell(row=cmp_row, column=3).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        for supplier_index, supplier in enumerate(compare_suppliers, start=4):
            price_pair = rec["prices"].get(supplier)
            if not price_pair:
                continue
            usd, cny = price_pair[0][type_index], price_pair[1][type_index]
            cell = ws_cmp.cell(row=cmp_row, column=supplier_index)
            if usd is None or cny is None:
                cell.value = "-"
            else:
                cell.value = f'=IF($B$2="USD",{fmt(usd)},{fmt(cny)})'
            cell.font = openpyxl.styles.Font(size=9)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.######"
        cmp_row += 1
    ws_cmp.merge_cells(start_row=block_start, start_column=1, end_row=cmp_row - 1, end_column=1)
    ws_cmp.merge_cells(start_row=block_start, start_column=2, end_row=cmp_row - 1, end_column=2)
    for c in (1, 2):
        ws_cmp.cell(row=block_start, column=c).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

cmp_last_col = 3 + len(compare_suppliers)
cmp_last_letter = openpyxl.utils.get_column_letter(cmp_last_col)
for row in ws_cmp.iter_rows(min_row=4, max_row=cmp_row - 1, min_col=1, max_col=cmp_last_col):
    for cell in row:
        cell.border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
ws_cmp.auto_filter.ref = f"A4:{cmp_last_letter}{cmp_row - 1}"
ws_cmp.column_dimensions["A"].width = 30
ws_cmp.column_dimensions["B"].width = 12
ws_cmp.column_dimensions["C"].width = 8
for j in range(len(compare_suppliers)):
    ws_cmp.column_dimensions[openpyxl.utils.get_column_letter(4 + j)].width = 13
ws_cmp.freeze_panes = "D5"

# ---------------------------------------------------------------- 渠道表（所有渠道/供应商 + 网址）
CH_HEADER = ["渠道(供应商)", "生产厂商", "数据来源", "模型数", "文档/定价网址", "API base",
             "汇率(CNY/计价单位)", "计价单位", "价格解析方式", "额外价格倍率"]
ch_rows = []
# models 侧渠道
for prov, pv in md.items():
    pven, pvzh = vendor_of(prov)
    vendor_label = pvzh if pven else "中转/云平台"
    channel_cfg = channel_settings(pv.get("name") or prov, prov)
    ch_rows.append([pv.get("name") or prov, vendor_label, "models.dev",
                    len(pv.get("models", {})), pricing_url_of(pv.get("name") or prov, pv.get("doc")), pv.get("api"),
                    channel_cfg["rate"], channel_cfg["unit"], channel_cfg.get("pricing_mode", "auto"), channel_cfg.get("price_scale", 1.0)])
for relay_channel in relaywatch_channels:
    ch_rows.append([
        relay_channel["name"], relay_channel["vendor"], relay_channel["source"], relay_channel["model_count"],
        relay_channel["pricing_url"], relay_channel["api"], relay_channel["rate"], relay_channel["unit"],
        relay_channel.get("pricing_mode", "auto"), relay_channel.get("price_scale", 1.0),
    ])
ws_ch = wb.create_sheet("渠道表")
ws_ch.append([safe(x) for x in CH_HEADER])
for r in ch_rows:
    ws_ch.append([safe(x) for x in r])
for c in range(1, len(CH_HEADER) + 1):
    cell = ws_ch.cell(row=1, column=c)
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D9E2F3")
    cell.border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws_ch.iter_rows(min_row=2, max_row=ws_ch.max_row, max_col=len(CH_HEADER)):
    for c in row:
        c.border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
ws_ch.column_dimensions["A"].width = 30
ws_ch.column_dimensions["B"].width = 16
ws_ch.column_dimensions["C"].width = 12
ws_ch.column_dimensions["D"].width = 10
ws_ch.column_dimensions["E"].width = 50
ws_ch.column_dimensions["F"].width = 42
ws_ch.column_dimensions["G"].width = 20
ws_ch.column_dimensions["H"].width = 12
ws_ch.column_dimensions["I"].width = 18
ws_ch.column_dimensions["J"].width = 16
ws_ch.freeze_panes = "A2"

wb.save(OUT)
print("已生成:", OUT)
print("models 表:", len(models_rows), "行 | 渠道表:", len(ch_rows), "行")
print("RelayWatch 模型行:", relaywatch_row_count, "| RelayWatch 渠道:", len(relaywatch_channels))
print("models 系列未命中(用fallback):", models_miss_fam, f"({models_miss_fam/len(models_rows):.1%})")
print("人工模型映射:", len(manual_model_map), "条")
print("models 有 doc 渠道:", sum(1 for pv in md.values() if pv.get("doc")), f"/{len(md)} | 有 api:", sum(1 for pv in md.values() if pv.get("api")), f"/{len(md)}")
