# -*- coding: utf-8 -*-
"""Convert the generated workbook into compact JSON for the local dashboard."""

from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_ROOT = ROOT
DEFAULT_WORKBOOK = APP_ROOT / "public" / "downloads" / "AI_API_models_审查版.xlsx"
OUTPUT_JSON = APP_ROOT / "public" / "data" / "dashboard.json"
INDEX_JSON = APP_ROOT / "public" / "data" / "index.json"
COMPARISON_DIR = APP_ROOT / "public" / "data" / "comparison"
BEST_DIR = APP_ROOT / "public" / "data" / "best"
DOWNLOAD_DIR = APP_ROOT / "public" / "downloads"
PAGE_SIZE = 24


def clean(value):
    return None if value in (None, "") else value


def to_number(value):
    value = clean(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def category_for_channel(name, vendor):
    if str(vendor or "").strip() == "中转/云平台":
        return "transit"
    if str(name or "").strip() == str(vendor or "").strip() and vendor:
        return "official"
    return "official" if vendor and vendor != "中转/云平台" else "unknown"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    args = parser.parse_args()
    workbook_path = Path(args.workbook).resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    models_sheet = wb["models"]
    model_rows = list(models_sheet.iter_rows(values_only=True))
    headers = [str(v or "") for v in model_rows[0]]
    ix = {name: headers.index(name) for name in headers}
    series_header = "模型系列" if "模型系列" in ix else "模型类型"
    input_header = "输入形式" if "输入形式" in ix else None
    refreshed_header = "价格最后刷新时间" if "价格最后刷新时间" in ix else None
    exchange_rate_header = "渠道汇率(CNY/计价单位)" if "渠道汇率(CNY/计价单位)" in ix else None
    billing_unit_header = "计价单位" if "计价单位" in ix else None
    pricing_mode_header = "价格解析方式" if "价格解析方式" in ix else None
    price_scale_header = "额外价格倍率" if "额外价格倍率" in ix else None
    usd_headers = ["输入价格USD/1M", "输出价格USD/1M", "缓存读取USD/1M", "缓存写入USD/1M"]
    cny_headers = ["输入价格CNY/1M", "输出价格CNY/1M", "缓存读取CNY/1M", "缓存写入CNY/1M"]
    types = ["输入", "输出", "缓存读", "缓存写"]

    model_map = {}
    supplier_set = set()
    vendor_set = set()
    model_series_set = set()
    for row in model_rows[1:]:
        model = clean(row[ix["模型名"]])
        supplier = clean(row[ix["供应商"]])
        if not model or not supplier:
            continue
        vendor = clean(row[ix["厂商"]]) or "其他"
        model_series = clean(row[ix[series_header]]) or model
        detail = {
            "model": str(model),
            "modelSeries": str(model_series),
            "vendor": str(vendor),
            "supplier": str(supplier),
            "originalModelName": str(clean(row[ix["原始模型名"]]) or model),
            "pricingUrl": str(clean(row[ix["供应商网址"]]) or ""),
            "mode": str(clean(row[ix["模式"]]) or ""),
            "inputModalities": str(clean(row[ix[input_header]]) or "") if input_header else "",
            "lastUpdated": str(clean(row[ix[refreshed_header]]) or "") if refreshed_header else "",
            "maxInputContext": clean(row[ix["最大输入上下文"]]),
            "maxOutput": clean(row[ix["最大输出"]]),
            "dataSource": str(clean(row[ix["数据来源"]]) or ""),
            "exchangeRate": (to_number(row[ix[exchange_rate_header]]) or 6.74) if exchange_rate_header else 6.74,
            "billingUnit": str(clean(row[ix[billing_unit_header]]) or "USD") if billing_unit_header else "USD",
            "pricingMode": str(clean(row[ix[pricing_mode_header]]) or "auto") if pricing_mode_header else "auto",
            "priceScale": (to_number(row[ix[price_scale_header]]) or 1) if price_scale_header else 1,
        }
        usd = [to_number(row[ix[label]]) for label in usd_headers]
        cny = [to_number(row[ix[label]]) for label in cny_headers]
        if all(v is None for v in usd) and all(v is None for v in cny):
            continue
        detail["usdPrices"] = usd
        detail["cnyPrices"] = cny
        supplier_set.add(str(supplier))
        vendor_set.add(str(vendor))
        model_series_set.add(str(model_series))
        record = model_map.setdefault(
            str(model),
            {"model": str(model), "modelSeries": str(model_series), "vendor": str(vendor), "prices": {}},
        )
        if not record.get("modelSeries") and model_series:
            record["modelSeries"] = str(model_series)
        record["prices"][str(supplier)] = {
            "usd": usd,
            "cny": cny,
            "details": detail,
        }

    suppliers = sorted(supplier_set, key=str.casefold)
    models = sorted(model_map.values(), key=lambda item: item["model"].casefold())
    comparison = []
    for record in models:
        for type_index, type_name in enumerate(types):
            prices = {}
            for supplier, value in record["prices"].items():
                usd = value["usd"][type_index]
                cny = value["cny"][type_index]
                if usd is None and cny is None:
                    continue
                prices[supplier] = {
                    "usd": usd,
                    "cny": cny,
                    "details": value.get("details", {}),
                }
            if prices:
                comparison.append(
                    {
                        "model": record["model"],
                        "modelSeries": record["modelSeries"],
                        "modelType": record["modelSeries"],
                        "vendor": record["vendor"],
                        "type": type_name,
                        "prices": prices,
                    }
                )

    channels = []
    seen_channels = set()
    if "渠道表" in wb.sheetnames:
        channels_sheet = wb["渠道表"]
        channel_rows = list(channels_sheet.iter_rows(values_only=True))
        for row in channel_rows[1:]:
            name = clean(row[0])
            if not name or str(name) in seen_channels:
                continue
            seen_channels.add(str(name))
            vendor = clean(row[1]) or "未标注"
            channel = {
                "id": str(name),
                "name": str(name),
                "vendor": str(vendor),
                "source": str(clean(row[2]) or "未标注"),
                "modelCount": int(row[3] or 0),
                "pricingUrl": str(clean(row[4]) or ""),
                "apiBase": str(clean(row[5]) or ""),
                "exchangeRate": (to_number(row[6]) or 6.74) if len(row) > 6 else 6.74,
                "billingUnit": str(clean(row[7]) or "USD") if len(row) > 7 else "USD",
                "pricingMode": str(clean(row[8]) or "auto") if len(row) > 8 else "auto",
                "priceScale": (to_number(row[9]) or 1) if len(row) > 9 else 1,
                "ingestMode": "relaywatch" if str(clean(row[2]) or "").strip().lower().startswith("relaywatch") else "manual",
                "category": category_for_channel(name, vendor),
            }
            channels.append(channel)

    channel_categories = {channel["name"]: channel["category"] for channel in channels}
    best_channels = []
    for record in models:
        candidates = []
        for supplier, price in record["prices"].items():
            usd = price["usd"]
            cny = price["cny"]
            primary = usd[0] if usd and usd[0] is not None else None
            output = usd[1] if usd and usd[1] is not None else None
            if primary is None and output is None:
                continue
            candidates.append(
                {
                    "supplier": supplier,
                    "category": channel_categories.get(supplier, "unknown"),
                    "prices": {"usd": usd, "cny": cny},
                    "details": price.get("details", {}),
                    "sortKey": primary if primary is not None else output,
                }
            )
        if candidates:
            best = min(candidates, key=lambda item: (item["sortKey"], item["prices"]["usd"][1] or 10**9))
            best_channels.append(
                {
                    "model": record["model"],
                    "modelSeries": record["modelSeries"],
                    "modelType": record["modelSeries"],
                    "vendor": record["vendor"],
                    "supplier": best["supplier"],
                    "category": best["category"],
                    "prices": best["prices"],
                    "details": best["details"],
                    "inputModalities": best["details"].get("inputModalities", ""),
                    "lastUpdated": best["details"].get("lastUpdated", ""),
                }
            )

    updated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    payload = {
        "meta": {
            "updatedAt": updated_at,
            "workbook": workbook_path.name,
            "modelCount": len(models),
            "supplierCount": len(suppliers),
            "vendorCount": len(vendor_set),
            "sourceRows": len(model_rows) - 1,
        },
        "filters": {
            "models": [item["model"] for item in models],
            "suppliers": suppliers,
            "vendors": sorted(vendor_set, key=str.casefold),
            "modelSeries": sorted(model_series_set, key=str.casefold),
            "modelTypes": sorted(model_series_set, key=str.casefold),
            "priceTypes": types,
            "types": types,
        },
        "comparison": comparison,
        "bestChannels": best_channels,
        "channels": channels,
    }
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    # Keep the browser's initial payload small. Each page contains only the
    # rows/cards for 24 models and is fetched from the local public directory
    # when that page becomes visible.
    for directory in (COMPARISON_DIR, BEST_DIR):
        directory.mkdir(parents=True, exist_ok=True)
        for stale_file in directory.glob("*.json"):
            stale_file.unlink()

    model_index = []
    comparison_files = []
    best_files = []
    for page_number, start in enumerate(range(0, len(models), PAGE_SIZE)):
        page_models = models[start : start + PAGE_SIZE]
        page_names = {item["model"] for item in page_models}
        page_rows = [row for row in comparison if row["model"] in page_names]
        page_cards = [card for card in best_channels if card["model"] in page_names]
        page_name = f"{page_number:04d}.json"
        comparison_path = COMPARISON_DIR / page_name
        best_path = BEST_DIR / page_name
        comparison_path.write_text(
            json.dumps({"rows": page_rows}, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        best_path.write_text(
            json.dumps({"cards": page_cards}, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        comparison_files.append(f"/data/comparison/{page_name}")
        best_files.append(f"/data/best/{page_name}")
        for item in page_models:
            model_index.append(
                {
                    "model": item["model"],
                    "modelSeries": item["modelSeries"],
                    "modelType": item["modelSeries"],
                    "vendor": item["vendor"],
                    "suppliers": sorted(item["prices"], key=str.casefold),
                    "page": page_number,
                }
            )

    index_payload = {
        "meta": payload["meta"],
        "filters": payload["filters"],
        "modelIndex": model_index,
        "channels": channels,
        "files": {"comparison": comparison_files, "best": best_files},
        "pageSize": PAGE_SIZE,
    }
    INDEX_JSON.write_text(
        json.dumps(index_payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    download_target = (DOWNLOAD_DIR / "AI_API_models_审查版.xlsx").resolve()
    if workbook_path != download_target:
        shutil.copy2(workbook_path, download_target)
    status_path = OUTPUT_JSON.parent / "status.json"
    status_path.write_text(
        json.dumps(
            {
                "state": "success",
                "finishedAt": updated_at,
                "message": "数据已准备",
                "source": workbook_path.name,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"models": len(models), "suppliers": len(suppliers), "channels": len(channels), "pages": len(comparison_files), "updatedAt": updated_at}, ensure_ascii=False))


if __name__ == "__main__":
    main()
